from build import *
from native import *
from hcomp import *
from wasm import *
import matplotlib.pyplot as plt
import os
import numpy as np
import openpyxl as xl
import scipy.stats as sp

N_TRIALS = 30

def main():
    build()

    tot_native_data = []

    for i in range(0, N_TRIALS):
        print("NATIVE BENCHMARK #{}:".format(i + 1))
        tot_native_data.append(native_bench())
        print()
    print("------------------")
    print()

    tot_hcomp_data = []

    for i in range(0, N_TRIALS):
        print("HCOMP BENCHMARK #{}:".format(i + 1))
        tot_hcomp_data.append(hcomp_bench())
        print()
    print("------------------")
    print()

    tot_wasm_data = []

    for i in range(0, N_TRIALS):
        print("WASM BENCHMARK #{}:".format(i + 1))
        tot_wasm_data.append(wasm_bench())
        print()
    print("------------------")
    print()

    #start plotting
    if not os.path.isdir("./dat"):
        os.mkdir("./dat/") #folder to store the dat

    # benchnam = getbenchmarks()
    benchnam = ["gemm.c.bin"]
    for b in benchnam:

        # avg_native = 0
        avg_hcomp = 0
        avg_wasm = 0
        # vals_native = []
        vals_hcomp = []
        vals_wasm = []

        # for m in tot_native_data:
        #     avg_native+=m[b]
        #     vals_native.append(m[b])
        for m in tot_hcomp_data:
            avg_hcomp+=m[b]
            vals_hcomp.append(m[b])
        for m in tot_wasm_data:
            avg_wasm+=m[b]
            vals_wasm.append(m[b])

        outpath = "./dat/{}/".format(b)

        os.system("rm -rf {}".format(outpath))
        os.mkdir(outpath)

        #make excel project
        wb = xl.Workbook()
        ws = wb.active
        ws.title = b

        # ws["A1"] = "Native"
        ws["C1"] = "HComp"
        ws["E1"] = "WASM"

        # for d in range(len(vals_native)):
        #     ws["A{}".format(d + 2)] = vals_native[d]

        for d in range(len(vals_hcomp)):
            ws["C{}".format(d + 2)] = vals_hcomp[d]

        for d in range(len(vals_wasm)):
            ws["E{}".format(d + 2)] = vals_wasm[d]

        ws["G1"] = "Averages:"
        # ws["G2"] = "Native:"
        # ws["H2"] = "=AVERAGE(A2:A{})".format(len(vals_native) + 1)
        ws["G3"] = "HComp:"
        ws["H3"] = "=AVERAGE(C2:C{})".format(len(vals_hcomp) + 1)
        ws["G4"] = "WASM:"
        ws["H4"] = "=AVERAGE(E2:E{})".format(len(vals_wasm) + 1)

        ws["I1"] = "Standard Deviations:"
        # ws["I2"] = "Native:"
        # ws["J2"] = "=STDEV(A2:A{})".format(len(vals_native) + 1)
        ws["I3"] = "HComp:"
        ws["J3"] = "=STDEV(C2:C{})".format(len(vals_hcomp) + 1)
        ws["I4"] = "WASM:"
        ws["J4"] = "=STDEV(E2:E{})".format(len(vals_wasm) + 1)

        dof = N_TRIALS - 1
        t_test = sp.ttest_ind(vals_hcomp, vals_wasm)
        print(t_test.pvalue)
        ws["G5"] = "P-Value:"
        ws["H5"] = t_test.pvalue

        wb.save("{}dat.xlsx".format(outpath))

        avg_hcomp/=len(tot_hcomp_data)
        avg_wasm/=len(tot_wasm_data)

        hcomp_stdev = np.std(np.array(vals_hcomp))
        wasm_stdev = np.std(np.array(vals_wasm))

        plt.figure(0)

        plt.figure(0)
        plt.bar(range(0, 2), [avg_hcomp, avg_wasm], color=["blue", "orange"], yerr=[hcomp_stdev, wasm_stdev])
        
        plt.xticks(range(0, 2), ["HComp", "WASM"])
        plt.ylabel("Execution speed in milliseconds")
        plt.title("Benchmarks")
        plt.savefig("{}graph.png".format(outpath))
        plt.close(0)


if __name__ == "__main__":
    main()